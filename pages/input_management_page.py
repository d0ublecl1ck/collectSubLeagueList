"""
输入管理页面 - 提供任务列表的查看、编辑和删除功能
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
from datetime import datetime
import re
import os
from openpyxl import Workbook

from .base_page import BasePage
from models import Task


class InputManagementPage(BasePage):
    """输入管理页面 - 任务管理界面"""
    
    def setup_ui(self):
        """设置输入管理页面的用户界面"""
        # 页面标题
        title_label = ttk.Label(
            self.frame, 
            text="任务管理", 
            font=('Arial', 16, 'bold')
        )
        title_label.pack(pady=(0, 20))
        
        # 工具栏
        toolbar_frame = ttk.Frame(self.frame)
        toolbar_frame.pack(fill=tk.X, pady=(0, 10))
        
        # 刷新按钮
        refresh_btn = ttk.Button(
            toolbar_frame,
            text="刷新列表",
            command=self.refresh_data,
            width=12
        )
        refresh_btn.pack(side=tk.LEFT, padx=(0, 10))
        
        # 删除按钮
        delete_btn = ttk.Button(
            toolbar_frame,
            text="删除选中",
            command=self.delete_selected,
            width=12
        )
        delete_btn.pack(side=tk.LEFT, padx=10)
        
        # 全选按钮
        select_all_btn = ttk.Button(
            toolbar_frame,
            text="全选",
            command=self.select_all,
            width=8
        )
        select_all_btn.pack(side=tk.LEFT, padx=(10, 5))
        
        # 反选按钮
        invert_btn = ttk.Button(
            toolbar_frame,
            text="反选",
            command=self.invert_selection,
            width=8
        )
        invert_btn.pack(side=tk.LEFT, padx=5)
        
        # 导出按钮
        export_btn = ttk.Button(
            toolbar_frame,
            text="导出Excel",
            command=self.export_to_excel,
            width=12
        )
        export_btn.pack(side=tk.LEFT, padx=(10, 0))
        
        # 生成其他年份按钮
        generate_year_btn = ttk.Button(
            toolbar_frame,
            text="生成其他年份",
            command=self.generate_other_year,
            width=12
        )
        generate_year_btn.pack(side=tk.LEFT, padx=(10, 0))
        
        # 删除年份按钮
        delete_year_btn = ttk.Button(
            toolbar_frame,
            text="删除年份",
            command=self.delete_year,
            width=12
        )
        delete_year_btn.pack(side=tk.LEFT, padx=(10, 0))
        
        # 搜索框
        search_frame = ttk.Frame(toolbar_frame)
        search_frame.pack(side=tk.RIGHT)
        
        ttk.Label(search_frame, text="搜索:").pack(side=tk.LEFT, padx=(0, 5))
        self.search_var = tk.StringVar()
        self.search_var.trace('w', self.on_search_change)
        search_entry = ttk.Entry(search_frame, textvariable=self.search_var, width=20)
        search_entry.pack(side=tk.LEFT)
        
        # 统计信息栏
        stats_frame = ttk.Frame(self.frame)
        stats_frame.pack(fill=tk.X, pady=(5, 0))
        
        self.stats_label = ttk.Label(
            stats_frame, 
            text="总数: 0 | 显示: 0 | 选中: 0",
            font=('Arial', 9),
            foreground='blue'
        )
        self.stats_label.pack(side=tk.LEFT)
        
        # 任务列表
        self.create_task_list()
        
        # 详情区域
        self.create_detail_panel()
        
        # 初始化完整项目列表（用于搜索状态管理）
        self.all_task_items = []
        
        # 加载数据
        self.refresh_data()
        
    def create_task_list(self):
        """创建任务列表"""
        list_frame = ttk.LabelFrame(self.frame, text="任务列表", padding=10)
        list_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 20))
        
        # 创建 Treeview
        columns = ('ID', '赛事级别', '赛事名称', '联赛名称', '国家', '年份', '类型', '分组', '主要链接', '备用链接', '最后爬取', '创建时间')
        self.task_tree = ttk.Treeview(list_frame, columns=columns, show='headings', height=12, selectmode='extended')
        
        # 设置列标题和宽度
        self.task_tree.heading('ID', text='ID')
        self.task_tree.heading('赛事级别', text='赛事级别')
        self.task_tree.heading('赛事名称', text='赛事名称')
        self.task_tree.heading('联赛名称', text='联赛名称')
        self.task_tree.heading('国家', text='国家')
        self.task_tree.heading('年份', text='年份')
        self.task_tree.heading('类型', text='类型')
        self.task_tree.heading('分组', text='分组')
        self.task_tree.heading('主要链接', text='主要链接')
        self.task_tree.heading('备用链接', text='备用链接')
        self.task_tree.heading('最后爬取', text='最后爬取')
        self.task_tree.heading('创建时间', text='创建时间')
        
        self.task_tree.column('ID', width=50, anchor='center')
        self.task_tree.column('赛事级别', width=80, anchor='center')
        self.task_tree.column('赛事名称', width=120)
        self.task_tree.column('联赛名称', width=150)
        self.task_tree.column('国家', width=80)
        self.task_tree.column('年份', width=60, anchor='center')
        self.task_tree.column('类型', width=80, anchor='center')
        self.task_tree.column('分组', width=80, anchor='center')
        self.task_tree.column('主要链接', width=150)
        self.task_tree.column('备用链接', width=150)
        self.task_tree.column('最后爬取', width=120, anchor='center')
        self.task_tree.column('创建时间', width=120, anchor='center')
        
        # 滚动条
        scrollbar_y = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=self.task_tree.yview)
        scrollbar_x = ttk.Scrollbar(list_frame, orient=tk.HORIZONTAL, command=self.task_tree.xview)
        self.task_tree.configure(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)
        
        # 布局
        self.task_tree.grid(row=0, column=0, sticky='nsew')
        scrollbar_y.grid(row=0, column=1, sticky='ns')
        scrollbar_x.grid(row=1, column=0, sticky='ew')
        
        list_frame.grid_rowconfigure(0, weight=1)
        list_frame.grid_columnconfigure(0, weight=1)
        
        # 绑定选择事件和双击编辑事件
        self.task_tree.bind('<<TreeviewSelect>>', self.on_task_select)
        self.task_tree.bind('<Double-1>', self.on_double_click)
        
        # 初始化编辑状态变量
        self.edit_item = None
        self.edit_column = None
        self.edit_widget = None
        
    def create_detail_panel(self):
        """创建详情面板"""
        detail_frame = ttk.LabelFrame(self.frame, text="任务详情", padding=15)
        detail_frame.pack(fill=tk.X)
        
        # 创建详情显示区域
        self.detail_text = tk.Text(
            detail_frame, 
            height=8, 
            wrap=tk.WORD,
            state=tk.DISABLED,
            font=('Consolas', 10)
        )
        
        detail_scrollbar = ttk.Scrollbar(detail_frame, orient=tk.VERTICAL, command=self.detail_text.yview)
        self.detail_text.configure(yscrollcommand=detail_scrollbar.set)
        
        self.detail_text.grid(row=0, column=0, sticky='nsew')
        detail_scrollbar.grid(row=0, column=1, sticky='ns')
        
        detail_frame.grid_rowconfigure(0, weight=1)
        detail_frame.grid_columnconfigure(0, weight=1)
        
    def refresh_data(self):
        """刷新任务数据"""
        try:
            # 清空现有数据和完整项目列表
            for item in self.task_tree.get_children():
                self.task_tree.delete(item)
            self.all_task_items = []
            
            # 从数据库加载数据
            with self.get_db_session() as session:
                tasks = session.query(Task).order_by(Task.created_at.desc()).all()
                
                for task in tasks:
                    created_time = task.created_at.strftime('%Y-%m-%d %H:%M') if task.created_at else ''
                    last_crawl_time = task.last_crawl_time.strftime('%Y-%m-%d %H:%M') if task.last_crawl_time else '从未爬取'
                    item_id = self.task_tree.insert('', 'end', values=(
                        task.id,
                        task.level,
                        task.event,
                        task.league,
                        task.country,
                        task.year,
                        task.type,
                        task.group,
                        task.link or '',
                        task.link_second or '',
                        last_crawl_time,
                        created_time
                    ))
                    # 保存到完整项目列表
                    self.all_task_items.append(item_id)
            
            self.log_action("刷新任务列表", f"加载了 {len(tasks)} 条记录")
            
            # 更新统计信息
            self.update_stats()
            
        except Exception as e:
            self.logger.error(f"刷新数据失败: {e}")
            self.show_message("错误", f"刷新数据失败: {str(e)}", "error")
    
    def on_task_select(self, event):
        """任务选择事件处理"""
        selection = self.task_tree.selection()
        if not selection:
            self.clear_detail()
            self.update_stats()
            return
            
        # 获取选中的任务ID
        item = self.task_tree.item(selection[0])
        task_id = item['values'][0]
        
        # 加载任务详情
        self.load_task_detail(task_id)
        
        # 更新统计信息
        self.update_stats()
    
    def load_task_detail(self, task_id):
        """加载任务详情"""
        try:
            with self.get_db_session() as session:
                task = session.query(Task).filter(Task.id == task_id).first()
                
                if task:
                    detail_text = f"""任务ID: {task.id}
赛事级别: {task.level}
赛事名称: {task.event}
国家/地区: {task.country}
联赛名称: {task.league}
赛事类型: {task.type}
赛事年份: {task.year}
分组信息: {task.group}
主要链接: {task.link or '未设置'}
备用链接: {task.link_second or '未设置'}
最后爬取: {task.last_crawl_time.strftime('%Y-%m-%d %H:%M:%S') if task.last_crawl_time else '从未爬取'}
创建时间: {task.created_at.strftime('%Y-%m-%d %H:%M:%S') if task.created_at else ''}
更新时间: {task.updated_at.strftime('%Y-%m-%d %H:%M:%S') if task.updated_at else ''}"""
                    
                    self.detail_text.config(state=tk.NORMAL)
                    self.detail_text.delete(1.0, tk.END)
                    self.detail_text.insert(1.0, detail_text)
                    self.detail_text.config(state=tk.DISABLED)
                    
        except Exception as e:
            self.logger.error(f"加载任务详情失败: {e}")
            self.show_message("错误", f"加载详情失败: {str(e)}", "error")
    
    def clear_detail(self):
        """清空详情显示"""
        self.detail_text.config(state=tk.NORMAL)
        self.detail_text.delete(1.0, tk.END)
        self.detail_text.config(state=tk.DISABLED)
    
    def delete_selected(self):
        """删除选中的任务"""
        selection = self.task_tree.selection()
        if not selection:
            self.show_message("提示", "请先选择要删除的任务", "warning")
            return
        
        # 构建待删除任务列表
        tasks_to_delete = []
        for item_id in selection:
            item = self.task_tree.item(item_id)
            task_id = item['values'][0]
            league_name = item['values'][3]
            tasks_to_delete.append((task_id, league_name))
        
        # 构建确认删除对话框消息
        if len(tasks_to_delete) == 1:
            task_id, league_name = tasks_to_delete[0]
            confirm_message = f"确定要删除任务 '{league_name}' (ID: {task_id}) 吗？\n\n此操作将同时删除相关的队伍数据，且无法恢复！"
        else:
            task_list = "\n".join([f"• {league} (ID: {tid})" for tid, league in tasks_to_delete])
            confirm_message = f"确定要删除以下 {len(tasks_to_delete)} 个任务吗？\n\n{task_list}\n\n此操作将同时删除相关的队伍数据，且无法恢复！"
        
        if not messagebox.askyesno("确认删除", confirm_message):
            return
        
        # 批量删除任务
        deleted_count = 0
        failed_tasks = []
        
        try:
            with self.get_db_session() as session:
                for task_id, league_name in tasks_to_delete:
                    try:
                        task = session.query(Task).filter(Task.id == task_id).first()
                        if task:
                            session.delete(task)
                            deleted_count += 1
                        else:
                            failed_tasks.append(f"{league_name} (ID: {task_id}) - 任务不存在")
                    except Exception as e:
                        failed_tasks.append(f"{league_name} (ID: {task_id}) - {str(e)}")
                
                session.commit()
                
                # 显示删除结果
                if failed_tasks:
                    result_message = f"删除完成！成功删除 {deleted_count} 个任务"
                    if failed_tasks:
                        result_message += f"\n\n失败的任务：\n" + "\n".join(failed_tasks)
                    self.show_message("删除结果", result_message, "warning")
                else:
                    self.show_message("成功", f"成功删除 {deleted_count} 个任务", "info")
                
                # 记录日志
                task_names = ", ".join([league for _, league in tasks_to_delete[:3]])
                if len(tasks_to_delete) > 3:
                    task_names += f" 等{len(tasks_to_delete)}个任务"
                self.log_action("批量删除任务", f"成功删除 {deleted_count} 个任务: {task_names}")
                
                # 刷新界面
                self.refresh_data()
                self.clear_detail()
                self.update_stats()
                    
        except Exception as e:
            self.logger.error(f"批量删除任务失败: {e}")
            self.show_message("错误", f"删除失败: {str(e)}", "error")
    
    def select_all(self):
        """全选所有可见任务"""
        try:
            # 获取所有可见项目（考虑搜索过滤）
            all_items = self.task_tree.get_children()
            if all_items:
                # 选中所有可见项目
                self.task_tree.selection_set(all_items)
                self.log_action("全选操作", f"已选中 {len(all_items)} 个任务")
                self.update_stats()
            else:
                self.show_message("提示", "暂无任务可选择", "info")
        except Exception as e:
            self.logger.error(f"全选操作失败: {e}")
            self.show_message("错误", f"全选失败: {str(e)}", "error")
    
    def invert_selection(self):
        """反选任务"""
        try:
            # 获取当前选中项目和所有可见项目
            current_selection = set(self.task_tree.selection())
            all_items = set(self.task_tree.get_children())
            
            if not all_items:
                self.show_message("提示", "暂无任务可操作", "info")
                return
            
            # 计算反选后的项目
            new_selection = all_items - current_selection
            
            # 应用新的选择
            self.task_tree.selection_set(list(new_selection))
            
            selected_count = len(new_selection)
            unselected_count = len(current_selection)
            self.log_action("反选操作", f"新选中 {selected_count} 个任务，取消选中 {unselected_count} 个任务")
            self.update_stats()
            
        except Exception as e:
            self.logger.error(f"反选操作失败: {e}")
            self.show_message("错误", f"反选失败: {str(e)}", "error")
    
    def on_search_change(self, *args):
        """搜索框内容变化事件"""
        search_text = self.search_var.get().lower()
        
        # 首先重新显示所有项目（解决搜索状态管理问题）
        for item in self.all_task_items:
            try:
                self.task_tree.reattach(item, '', 'end')
            except tk.TclError:
                # 如果项目已经被删除或不存在，跳过
                pass
        
        # 如果搜索框为空，显示所有项目
        if not search_text:
            self.update_stats()
            return
        
        # 过滤显示匹配的项目，使用完整项目列表确保所有数据都参与搜索
        for item in self.all_task_items:
            try:
                values = self.task_tree.item(item)['values']
                # 在所有字段中搜索：ID、赛事级别、赛事名称、联赛名称、国家、年份、类型、分组、最后爬取、创建时间
                if any(search_text in str(value).lower() for value in values):
                    self.task_tree.reattach(item, '', 'end')
                else:
                    self.task_tree.detach(item)
            except tk.TclError:
                # 如果项目已经被删除或不存在，跳过
                pass
        
        # 更新统计信息
        self.update_stats()
    
    def on_double_click(self, event):
        """处理双击事件"""
        # 取消当前编辑（如果有）
        self.cancel_edit()
        
        # 确定双击的位置
        item = self.task_tree.identify_row(event.y)
        column = self.task_tree.identify_column(event.x)
        
        if not item or not column:
            return
        
        # 获取列名
        column_index = int(column[1:]) - 1  # 列编号从#1开始
        columns = ('ID', '赛事级别', '赛事名称', '联赛名称', '国家', '年份', '类型', '分组', '主要链接', '备用链接', '最后爬取', '创建时间')
        
        if column_index < 0 or column_index >= len(columns):
            return
            
        column_name = columns[column_index]
        
        # 检查是否可编辑
        readonly_columns = ['ID', '最后爬取', '创建时间']
        if column_name in readonly_columns:
            self.show_message("提示", f"{column_name} 字段不可编辑", "warning")
            return
        
        # 开始编辑
        self.start_edit_cell(item, column_name, column_index)
    
    def start_edit_cell(self, item, column_name, column_index):
        """开始编辑单元格"""
        # 记录编辑状态
        self.edit_item = item
        self.edit_column = column_name
        
        # 获取当前值
        current_values = self.task_tree.item(item, 'values')
        current_value = current_values[column_index]
        
        # 获取单元格位置
        bbox = self.task_tree.bbox(item, f'#{column_index + 1}')
        if not bbox:
            return
        
        x, y, width, height = bbox
        
        # 创建编辑控件
        if column_name == '类型':
            # 下拉框编辑
            self.edit_widget = ttk.Combobox(
                self.task_tree,
                values=['常规', '东西拆分', '联二合并', '春秋合并'],
                state='readonly'
            )
            self.edit_widget.set(current_value)
        else:
            # 文本框编辑
            self.edit_widget = tk.Entry(self.task_tree)
            self.edit_widget.insert(0, current_value)
            self.edit_widget.select_range(0, tk.END)
        
        # 放置编辑控件
        self.edit_widget.place(x=x, y=y, width=width, height=height)
        self.edit_widget.focus_set()
        
        # 绑定事件
        self.edit_widget.bind('<Return>', self.save_edit)
        self.edit_widget.bind('<Escape>', self.cancel_edit)
        self.edit_widget.bind('<FocusOut>', self.save_edit)
    
    def save_edit(self, event=None):
        """保存编辑"""
        if not self.edit_widget or not self.edit_item:
            return
        
        try:
            # 获取新值
            new_value = self.edit_widget.get().strip()
            
            # 验证输入
            if not self.validate_input(self.edit_column, new_value):
                return
            
            # 获取任务ID
            item_values = self.task_tree.item(self.edit_item, 'values')
            task_id = item_values[0]
            
            # 更新数据库
            with self.get_db_session() as session:
                task = session.query(Task).filter(Task.id == task_id).first()
                if task:
                    # 根据列名更新对应字段
                    field_mapping = {
                        '赛事级别': 'level',
                        '赛事名称': 'event',
                        '联赛名称': 'league',
                        '国家': 'country',
                        '年份': 'year',
                        '类型': 'type',
                        '分组': 'group',
                        '主要链接': 'link',
                        '备用链接': 'link_second'
                    }
                    
                    if self.edit_column in field_mapping:
                        field_name = field_mapping[self.edit_column]
                        setattr(task, field_name, new_value if new_value else None)
                        task.updated_at = datetime.now()
                        session.commit()
                        
                        # 刷新显示
                        self.refresh_data()
                        self.log_action("编辑任务", f"更新任务 {task_id} 的 {self.edit_column}: {new_value}")
                        self.update_stats()
            
        except Exception as e:
            self.logger.error(f"保存编辑失败: {e}")
            self.show_message("错误", f"保存失败: {str(e)}", "error")
        finally:
            self.cancel_edit()
    
    def cancel_edit(self, event=None):
        """取消编辑"""
        if self.edit_widget:
            self.edit_widget.destroy()
            self.edit_widget = None
        self.edit_item = None
        self.edit_column = None
    
    def validate_input(self, column_name, value):
        """验证输入"""
        if column_name in ['赛事级别', '赛事名称', '联赛名称', '国家', '分组'] and not value:
            self.show_message("错误", f"{column_name} 不能为空", "error")
            return False
        
        if column_name == '赛事级别':
            try:
                int(value)
            except ValueError:
                self.show_message("错误", "赛事级别必须为数字", "error")
                return False
        
        if column_name == '年份':
            if value:
                try:
                    year = int(value)
                    if len(value) != 4 or year < 1900 or year > 2100:
                        self.show_message("错误", "年份格式不正确（请输入4位数字，如2024）", "error")
                        return False
                except ValueError:
                    self.show_message("错误", "年份必须为数字", "error")
                    return False
        
        if column_name == '类型':
            valid_types = ['常规', '东西拆分', '联二合并', '春秋合并']
            if value not in valid_types:
                self.show_message("错误", f"类型必须为: {'/'.join(valid_types)}", "error")
                return False
        
        if column_name in ['主要链接', '备用链接'] and value:
            url_pattern = re.compile(
                r'^https?://'  # http:// or https://
                r'(?:(?:[A-Z0-9](?:[A-Z0-9-]{0,61}[A-Z0-9])?\.)+[A-Z]{2,6}\.?|'  # domain...
                r'localhost|'  # localhost...
                r'\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})'  # ...or ip
                r'(?::\d+)?'  # optional port
                r'(?:/?|[/?]\S+)$', re.IGNORECASE)
            
            if not url_pattern.match(value):
                self.show_message("错误", f"{column_name} URL格式不正确", "error")
                return False
        
        return True
    
    def export_to_excel(self):
        """导出任务数据到Excel文件"""
        try:
            # 从数据库查询所有任务数据
            with self.get_db_session() as session:
                tasks = session.query(Task).order_by(Task.created_at.desc()).all()
                
                if not tasks:
                    self.show_message("提示", "暂无任务数据可导出", "info")
                    return
                
                # 让用户选择保存位置
                file_path = filedialog.asksaveasfilename(
                    title="保存Excel文件",
                    defaultextension=".xlsx",
                    filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")],
                    initialfile="任务数据导出.xlsx"
                )
                
                if not file_path:
                    return  # 用户取消了保存
                
                # 创建工作簿和工作表
                wb = Workbook()
                ws = wb.active
                ws.title = "任务数据"
                
                # 设置表头
                headers = ['赛事级别', '赛事名称', '国家/地区', '联赛名称', '赛事类型', '主要链接', '第二链接', '赛事年份', '分组信息']
                ws.append(headers)
                
                # 添加数据行
                for task in tasks:
                    row_data = [
                        task.level,
                        task.event or '',
                        task.country or '',
                        task.league or '',
                        task.type or '',
                        task.link or '',
                        task.link_second or '',
                        task.year or '',
                        task.group or ''
                    ]
                    ws.append(row_data)
                
                # 调整列宽
                column_widths = [10, 20, 15, 25, 12, 40, 40, 12, 15]
                for i, width in enumerate(column_widths, 1):
                    ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width
                
                # 设置表头样式
                for cell in ws[1]:
                    cell.font = cell.font.copy(bold=True)
                
                # 保存文件
                wb.save(file_path)
                
                # 显示成功消息
                self.show_message("成功", f"成功导出 {len(tasks)} 条任务数据到:\n{os.path.basename(file_path)}", "info")
                self.log_action("导出Excel", f"成功导出 {len(tasks)} 条任务数据到 {file_path}")
                
        except Exception as e:
            self.logger.error(f"导出Excel失败: {e}")
            self.show_message("错误", f"导出失败: {str(e)}", "error")
    
    def update_stats(self):
        """更新统计信息显示"""
        try:
            # 获取数据库总数
            with self.get_db_session() as session:
                total_count = session.query(Task).count()
            
            # 获取当前显示数量
            visible_count = len(self.task_tree.get_children())
            
            # 获取选中数量
            selected_count = len(self.task_tree.selection())
            
            # 更新显示
            stats_text = f"总数: {total_count} | 显示: {visible_count} | 选中: {selected_count}"
            self.stats_label.config(text=stats_text)
            
        except Exception as e:
            self.logger.error(f"更新统计信息失败: {e}")
            self.stats_label.config(text="统计信息加载失败")
    
    def generate_other_year(self):
        """生成其他年份的任务数据"""
        try:
            # 显示年份输入对话框
            year_input = simpledialog.askstring(
                "生成其他年份",
                "请输入目标年份 (如: 2022)\n"
                "系统将自动生成跨年格式 (如: 2022-2023)\n"
                "并替换链接中对应的年份",
                initialvalue=""
            )
            
            if not year_input:
                return  # 用户取消了输入
            
            # 验证年份输入
            is_valid, error_msg = self.validate_year_input(year_input)
            if not is_valid:
                self.show_message("输入错误", error_msg, "error")
                return
            
            target_year = int(year_input)
            
            # 确认操作
            confirm_msg = (f"确定要为年份 {target_year} 生成新的任务数据吗？\n\n"
                          f"系统将基于现有所有任务生成对应的 {target_year} 年份数据\n"
                          f"跨年格式将生成为 {target_year}-{target_year + 1}")
            
            if not messagebox.askyesno("确认生成", confirm_msg):
                return
            
            # 执行批量生成
            success_count, skip_count, error_count = self.generate_tasks_for_year(target_year)
            
            # 显示结果
            result_msg = (f"年份 {target_year} 数据生成完成！\n\n"
                         f"✓ 成功生成: {success_count} 条新任务\n"
                         f"⚠ 已存在跳过: {skip_count} 条任务\n"
                         f"✗ 处理失败: {error_count} 条任务")
            
            if success_count > 0:
                self.show_message("生成成功", result_msg, "info")
                # 刷新界面显示
                self.refresh_data()
                self.log_action("生成其他年份", f"为年份 {target_year} 成功生成 {success_count} 条新任务")
            else:
                self.show_message("生成结果", result_msg, "warning")
                
        except Exception as e:
            self.logger.error(f"生成其他年份失败: {e}")
            self.show_message("错误", f"生成失败: {str(e)}", "error")
    
    def validate_year_input(self, year_str):
        """验证年份输入"""
        if not year_str or not year_str.strip():
            return False, "年份不能为空"
        
        year_str = year_str.strip()
        
        if not year_str.isdigit():
            return False, "年份必须为数字"
        
        if len(year_str) != 4:
            return False, "年份必须为4位数字（如：2022）"
        
        year = int(year_str)
        if year < 1900 or year > 2100:
            return False, "年份范围必须在1900-2100之间"
        
        return True, ""
    
    def transform_year_field(self, original_year, target_year):
        """转换year字段的年份"""
        if not original_year:
            return str(target_year)
        
        original_str = str(original_year).strip()
        
        # 匹配单年格式 "2023"
        if re.match(r'^\d{4}$', original_str):
            return str(target_year)
        
        # 匹配跨年格式 "2022-2023"
        elif re.match(r'^\d{4}-\d{4}$', original_str):
            return f"{target_year}-{target_year + 1}"
        
        # 其他格式保持不变
        else:
            return original_str
    
    def transform_url_year(self, original_url, target_year):
        """转换URL中的年份"""
        if not original_url or not original_url.strip():
            return original_url
        
        url = original_url.strip()
        
        # 匹配单年格式：/2023/
        pattern1 = r'/(\d{4})/'
        if re.search(pattern1, url):
            return re.sub(pattern1, f'/{target_year}/', url)
        
        # 匹配跨年格式：/2022-2023/
        pattern2 = r'/(\d{4})-(\d{4})/'
        if re.search(pattern2, url):
            return re.sub(pattern2, f'/{target_year}-{target_year + 1}/', url)
        
        # 无匹配则保持不变
        return url
    
    def check_task_exists(self, session, league, year, group):
        """检查任务是否已存在（基于uk_league_year_group约束）"""
        existing = session.query(Task).filter(
            Task.league == league,
            Task.year == year,
            Task.group == group
        ).first()
        return existing is not None
    
    def generate_tasks_for_year(self, target_year):
        """为指定年份生成任务"""
        success_count = 0
        skip_count = 0
        error_count = 0
        
        try:
            with self.get_db_session() as session:
                # 读取所有现有任务
                original_tasks = session.query(Task).all()
                
                if not original_tasks:
                    return 0, 0, 0
                
                for original_task in original_tasks:
                    try:
                        # 生成新任务数据
                        new_year = self.transform_year_field(original_task.year, target_year)
                        new_link = self.transform_url_year(original_task.link, target_year)
                        new_link_second = self.transform_url_year(original_task.link_second, target_year)
                        
                        # 检查是否已存在（基于uk_league_year_group约束）
                        if self.check_task_exists(session, original_task.league, new_year, original_task.group):
                            skip_count += 1
                            continue
                        
                        # 创建新任务
                        new_task = Task(
                            level=original_task.level,
                            event=original_task.event,
                            country=original_task.country,
                            league=original_task.league,
                            type=original_task.type,
                            year=new_year,
                            group=original_task.group,
                            link=new_link,
                            link_second=new_link_second
                        )
                        
                        session.add(new_task)
                        success_count += 1
                        
                    except Exception as e:
                        error_count += 1
                        self.logger.error(f"处理任务 {original_task.id} 失败: {e}")
                
                # 提交所有更改
                session.commit()
                
        except Exception as e:
            self.logger.error(f"批量生成任务失败: {e}")
            raise
        
        return success_count, skip_count, error_count
    
    def delete_year(self):
        """删除指定年份的所有任务数据"""
        try:
            # 显示年份输入对话框
            year_input = simpledialog.askstring(
                "删除年份",
                "请输入要删除的年份 (如: 2024)\n\n"
                "系统将删除该年份的所有任务数据\n"
                "包括单年格式 (2024) 和跨年格式 (2024-2025)\n\n"
                "⚠️ 警告：此操作不可恢复！",
                initialvalue=""
            )
            
            if not year_input:
                return  # 用户取消了输入
            
            # 验证年份输入
            is_valid, error_msg = self.validate_year_input(year_input)
            if not is_valid:
                self.show_message("输入错误", error_msg, "error")
                return
            
            target_year = int(year_input)
            
            # 查询要删除的任务
            tasks_to_delete = self.find_tasks_by_year(target_year)
            
            if not tasks_to_delete:
                self.show_message("提示", f"未找到年份为 {target_year} 的任务数据", "info")
                return
            
            # 显示删除明细并确认
            if self.confirm_year_deletion(target_year, tasks_to_delete):
                # 执行删除操作
                self.execute_year_deletion(target_year, tasks_to_delete)
                
        except Exception as e:
            self.logger.error(f"删除年份失败: {e}")
            self.show_message("错误", f"删除失败: {str(e)}", "error")
    
    def find_tasks_by_year(self, target_year):
        """查找指定年份的所有任务"""
        tasks_to_delete = []
        
        try:
            with self.get_db_session() as session:
                # 构建年份匹配条件：单年格式和跨年格式
                single_year = str(target_year)
                cross_year = f"{target_year}-{target_year + 1}"
                
                # 使用SQL查询直接筛选匹配的年份，避免查询所有任务后再过滤
                matching_tasks = session.query(Task).filter(
                    (Task.year == single_year) | (Task.year == cross_year)
                ).all()
                
                # 构建删除列表
                for task in matching_tasks:
                    tasks_to_delete.append({
                        'id': task.id,
                        'league': task.league,
                        'country': task.country,
                        'year': task.year,
                        'type': task.type,
                        'created_at': task.created_at
                    })
                        
        except Exception as e:
            self.logger.error(f"查询年份任务失败: {e}")
            raise
        
        return tasks_to_delete
    
    def confirm_year_deletion(self, target_year, tasks_to_delete):
        """显示删除明细并确认操作"""
        # 构建明细信息
        detail_lines = []
        detail_lines.append(f"找到 {len(tasks_to_delete)} 条匹配的任务数据\n")
        
        # 按类型分组统计
        single_year_count = 0
        cross_year_count = 0
        type_stats = {}
        country_stats = {}
        
        for task in tasks_to_delete:
            # 年份格式统计
            if str(task['year']) == str(target_year):
                single_year_count += 1
            elif str(task['year']) == f"{target_year}-{target_year + 1}":
                cross_year_count += 1
            
            # 类型统计
            task_type = task['type'] or '未设置'
            type_stats[task_type] = type_stats.get(task_type, 0) + 1
            
            # 国家统计
            country = task['country'] or '未设置'
            country_stats[country] = country_stats.get(country, 0) + 1
        
        # 添加统计信息（紧凑显示）
        detail_lines.append("📊 删除统计：")
        detail_lines.append(f"  单年 ({target_year}): {single_year_count} | 跨年 ({target_year}-{target_year + 1}): {cross_year_count}")
        
        # 类型统计（限制显示数量）
        if type_stats:
            type_items = sorted(type_stats.items())[:3]  # 只显示前3种类型
            type_summary = " | ".join([f"{t}: {c}" for t, c in type_items])
            detail_lines.append(f"🏆 主要类型: {type_summary}")
            if len(type_stats) > 3:
                detail_lines.append(f"  还有 {len(type_stats) - 3} 种其他类型")
        
        # 国家统计（限制显示数量）
        if country_stats:
            country_items = sorted(country_stats.items(), key=lambda x: x[1], reverse=True)[:3]  # 按数量排序，显示前3个
            country_summary = " | ".join([f"{c}: {n}" for c, n in country_items])
            detail_lines.append(f"🌍 主要国家: {country_summary}")
            if len(country_stats) > 3:
                detail_lines.append(f"  还有 {len(country_stats) - 3} 个其他国家")
        
        # 只显示前5条具体任务，避免对话框过长
        detail_lines.append("\n📝 部分任务明细:")
        for i, task in enumerate(tasks_to_delete[:5]):
            detail_lines.append(f"  {i+1}. [{task['year']}] {task['country']} - {task['league']}")
        
        if len(tasks_to_delete) > 5:
            detail_lines.append(f"  ... 还有 {len(tasks_to_delete) - 5} 条任务")
        
        detail_lines.append("\n⚠️ 警告：此操作不可恢复！")
        
        confirm_message = "\n".join(detail_lines)
        
        # 显示确认对话框
        return messagebox.askyesno(
            f"确认删除年份 {target_year}",
            confirm_message
        )
    
    def execute_year_deletion(self, target_year, tasks_to_delete):
        """执行年份删除操作"""
        if not tasks_to_delete:
            return
            
        try:
            with self.get_db_session() as session:
                # 提取所有要删除的任务ID
                task_ids = [task_info['id'] for task_info in tasks_to_delete]
                
                # 使用批量删除SQL语句，避免页面阻塞
                # 这里使用SQLAlchemy的bulk delete功能
                deleted_count = session.query(Task).filter(Task.id.in_(task_ids)).delete(synchronize_session=False)
                
                # 提交删除操作
                session.commit()
                
                # 显示结果
                expected_count = len(tasks_to_delete)
                result_msg = f"年份 {target_year} 批量删除操作完成！\n\n"
                result_msg += f"✓ 成功删除: {deleted_count} 条任务\n"
                
                # 检查是否有未删除的任务
                if deleted_count < expected_count:
                    failed_count = expected_count - deleted_count
                    result_msg += f"⚠ 未删除: {failed_count} 条任务（可能已被删除或不存在）\n"
                
                if deleted_count > 0:
                    self.show_message("删除成功", result_msg, "info")
                    # 记录日志
                    self.log_action("批量删除年份", f"成功批量删除年份 {target_year} 的 {deleted_count} 条任务")
                    # 刷新界面
                    self.refresh_data()
                    self.clear_detail()
                    self.update_stats()
                else:
                    self.show_message("删除结果", "未删除任何任务，可能目标数据已不存在", "warning")
                    
        except Exception as e:
            self.logger.error(f"执行批量年份删除失败: {e}")
            self.show_message("错误", f"批量删除操作失败: {str(e)}", "error")